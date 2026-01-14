"""
ExcelHelper - Encapsulates all Excel/SharePoint operations
No need to pass config/logger to every method - stored in instance
"""
import json
import os
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any

logger = logging.getLogger(__name__)


class ExcelHelper:
    """Handles all Excel and SharePoint Excel operations"""
    
    def __init__(self, config: dict, transport_config: dict, get_access_token_func, 
                 logger_instance=None, app_logger_instance=None):
        """
        Initialize Excel helper with configuration and dependencies
        
        Args:
            config: Full configuration dict
            transport_config: Transport section of config
            get_access_token_func: Function to get SharePoint access token
            logger_instance: Standard logger
            app_logger_instance: Structured app logger
        """
        self.config = config
        self.transport_config = transport_config
        self.sharepoint_config = transport_config.get('sharepoint', {})
        self.get_access_token = get_access_token_func
        self.logger = logger_instance or logger
        self.app_logger = app_logger_instance
        
    def save_to_excel(self, request_id: str, data: dict, has_attachment: bool = False, 
                     attachment_error: str = None, json_index: int = None,
                     update_json_sync_status_func=None, attachment_status: str = None) -> Dict[str, Any]:
        """
        Save transport request to Excel/SharePoint
        
        Args:
            request_id: Request identifier
            data: Form data
            has_attachment: Whether request has attachments (for bool logic)
            attachment_error: Error message if attachment failed
            json_index: Index in JSON backup (for sync status update)
            update_json_sync_status_func: Function to update JSON sync status
            attachment_status: Explicit status override ('Processing', 'Saved', 'Failed', 'None')
            
        Returns:
            dict: Result with sharepoint_saved, sharepoint_error, json_index keys
        """
        if self.app_logger:
            self.app_logger.log_info(
                f"\033[94mℹ save_to_excel() STARTED for request: {request_id}\033[0m",
                {'function': 'save_to_excel', 'request_id': request_id}
            )
        
        result = {
            'local_saved': False,
            'sharepoint_saved': False,
            'sharepoint_error': None,
            'local_error': None,
            'json_index': json_index
        }
        
        # Prepare row data
        row_data = {
            'Request_ID': request_id,
            'Timestamp': datetime.now().isoformat(),
            'Delivery_Note_Number': data.get('deliveryNoteNumber', ''),
            'Truck_License_Plates': data.get('truckLicensePlates', ''),
            'Trailer_License_Plates': data.get('trailerLicensePlates', ''),
            'Carrier_Country': data.get('carrierCountry', ''),
            'Carrier_Tax_Code': data.get('carrierTaxCode', ''),
            'Carrier_Full_Name': data.get('carrierFullName', ''),
            'Border_Crossing': data.get('borderCrossing', ''),
            'Border_Crossing_Date': data.get('borderCrossingDate', ''),
            'Email': data.get('email', ''),
            'Phone_Number': data.get('phoneNumber', ''),
            'Has_Attachment': attachment_status if attachment_status else ('Yes' if has_attachment else 'No'),
            'Attachment_Status': attachment_status if attachment_status else ('Saved' if has_attachment else ('Failed' if attachment_error else 'None')),
            'Attachment_Error': attachment_error if attachment_error else '',
            'SharePoint_Synced': False
        }
        
        # JSON backup is already saved in main endpoint
        result['local_saved'] = True
        
        # SharePoint Integration
        if self.sharepoint_config.get('enabled', False):
            try:
                use_excel_api = self.sharepoint_config.get('use_excel_api', True)
                
                if use_excel_api:
                    # Use Excel API (direct row insertion)
                    self._save_via_excel_api(request_id, data, has_attachment, attachment_error)
                    self.logger.info("\033[92m✓ Data saved to SharePoint Excel via API\033[0m")
                else:
                    # Use traditional download/upload method
                    max_retries = self.sharepoint_config.get('max_retries', 3)
                    retry_wait_multiplier = self.sharepoint_config.get('retry_wait_multiplier', 2)
                    self._save_via_traditional(
                        request_id, data, has_attachment, attachment_error,
                        max_retries, retry_wait_multiplier
                    )
                    self.logger.info("\033[92m✓ Data saved to SharePoint Excel\033[0m")
                
                result['sharepoint_saved'] = True
                
                # Update JSON sync status
                if result.get('local_saved') and json_index is not None and update_json_sync_status_func:
                    try:
                        update_json_sync_status_func(request_id, json_index, synced=True)
                    except Exception as update_error:
                        self.logger.warning(f"Failed to update JSON sync status: {update_error}")
                        
            except Exception as sp_error:
                error_msg = str(sp_error)
                self.logger.error(f"\033[91m✗ SharePoint save failed: {error_msg}\033[0m", exc_info=True)
                result['sharepoint_error'] = error_msg
        
        if self.app_logger:
            self.app_logger.log_info(
                f"\033[94mℹ save_to_excel() COMPLETED - Result: {result}\033[0m",
                {'function': 'save_to_excel', 'result': result}
            )
        
        return result
    
    def _save_via_excel_api(self, request_id: str, data: dict, 
                           has_attachment: bool, attachment_error: str):
        """
        Save record using Excel API (direct row insertion)
        Works even when file is open by other users
        """
        try:
            from sharepoint_helper import SharePointHelper
            
            access_token = self.get_access_token()
            sp = SharePointHelper(access_token)
            
            folder_url = self.sharepoint_config['folder_url']
            excel_file_name = self.sharepoint_config['excel_file_name']
            worksheet_name = self.sharepoint_config.get('worksheet_name', 'Sheet1')
            
            self.logger.info(f"SharePoint Excel API: Adding row to {excel_file_name}")
            
            # Prepare row values in correct column order
            attachment_status = 'Saved' if has_attachment else ('Failed' if attachment_error else 'None')
            
            row_values = [
                request_id,
                data.get('deliveryNoteNumber', ''),
                data.get('truckLicensePlates', ''),
                data.get('trailerLicensePlates', ''),
                data.get('carrierCountry', ''),
                data.get('carrierTaxCode', ''),
                data.get('carrierFullName', ''),
                data.get('borderCrossing', ''),
                data.get('borderCrossingDate', ''),
                data.get('email', ''),
                data.get('phoneNumber', ''),
                'Yes' if has_attachment else 'No',
                attachment_status,
                attachment_error if attachment_error else ''
            ]
            
            # Add row using Excel API
            sp.add_excel_row(
                folder_url=folder_url,
                excel_file_name=excel_file_name,
                worksheet_name=worksheet_name,
                row_values=row_values
            )
            
            self.logger.info(f"\033[92m✓ SharePoint Excel API: Successfully added row for {request_id}\033[0m")
            
        except Exception as e:
            self.logger.error(f"\033[91m✗ SharePoint Excel API error: {e}\033[0m", exc_info=True)
            raise
    
    def _save_via_traditional(self, request_id: str, data: dict,
                             has_attachment: bool, attachment_error: str,
                             max_retries: int, retry_wait_multiplier: int):
        """
        Save record using traditional download/upload method
        Uses retry logic for locked files
        """
        import time
        import tempfile
        from pathlib import Path
        from datetime import datetime
        from sharepoint_helper import SharePointHelper
        from openpyxl import load_workbook
        
        access_token = self.get_access_token()
        sp = SharePointHelper(access_token)
        
        folder_url = self.sharepoint_config['folder_url']
        excel_file_name = self.sharepoint_config['excel_file_name']
        worksheet_name = self.sharepoint_config.get('worksheet_name', 'Sheet1')
        
        self.logger.info(f"SharePoint: Getting folder from {folder_url}")
        folder = sp.get_folder(folder_url)
        
        # Retry loop for handling locked files
        last_error = None
        for attempt in range(max_retries):
            try:
                with tempfile.TemporaryDirectory() as temp_dir:
                    temp_path = Path(temp_dir) / excel_file_name
                    
                    self.logger.info(f"SharePoint: Downloading {excel_file_name} (attempt {attempt + 1}/{max_retries})")
                    sp.download_file_from_folder(
                        download_path=temp_dir,
                        folder=folder,
                        file_name=excel_file_name
                    )
                    
                    wb = load_workbook(temp_path)
                    ws = wb[worksheet_name] if worksheet_name in wb.sheetnames else wb.active
                    
                    # Prepare new row with data
                    new_row = [
                        request_id,
                        datetime.now().isoformat(),
                        data.get('deliveryNoteNumber', ''),
                        data.get('truckLicensePlates', ''),
                        data.get('trailerLicensePlates', ''),
                        data.get('carrierCountry', ''),
                        data.get('carrierTaxCode', ''),
                        data.get('carrierFullName', ''),
                        data.get('borderCrossing', ''),
                        data.get('borderCrossingDate', ''),
                        data.get('email', ''),
                        data.get('phoneNumber', ''),
                        'Yes' if has_attachment else 'No',
                        'Saved' if has_attachment else ('Failed' if attachment_error else 'None'),
                        attachment_error if attachment_error else ''
                    ]
                    
                    ws.append(new_row)
                    self.logger.info(f"SharePoint: Added new row {ws.max_row}")
                    
                    wb.save(temp_path)
                    wb.close()
                    
                    self.logger.info(f"SharePoint: Uploading {excel_file_name} (attempt {attempt + 1}/{max_retries})")
                    sp.upload_file(temp_path, folder)
                    
                    self.logger.info(f"\033[92m✓ SharePoint: Successfully saved to {excel_file_name}\033[0m")
                    return
                    
            except Exception as e:
                last_error = e
                error_str = str(e).lower()
                
                is_locked = any(indicator in error_str for indicator in [
                    'locked', 'in use', 'being used', 'cannot access', '423',
                    'cobaltlockviolation', 'resourcelocked', 'file is open'
                ])
                
                if is_locked and attempt < max_retries - 1:
                    wait_time = (attempt + 1) * retry_wait_multiplier
                    self.logger.warning(f"\033[93m⚠ SharePoint: File locked. Waiting {wait_time}s before retry...\033[0m")
                    time.sleep(wait_time)
                    continue
                else:
                    if is_locked:
                        self.logger.error(f"\033[91m✗ SharePoint: File still locked after {max_retries} attempts\033[0m")
                    raise
        
        self.logger.error(f"\033[91m✗ SharePoint: All {max_retries} attempts failed\033[0m")
        raise Exception(f"Failed to save to SharePoint after {max_retries} attempts. Last error: {last_error}")

    def update_attachment_status(self, request_id: str, has_attachment: bool, 
                                attachment_error: str = None) -> Dict[str, Any]:
        """
        Update Excel with final attachment status
        
        Args:
            request_id: Request identifier
            has_attachment: Whether attachments were saved
            attachment_error: Error message if failed
            
        Returns:
            dict: Result with success status
        """
        try:
            from sharepoint_helper import SharePointHelper
            
            # Get access token
            access_token = self.get_access_token()
            sp = SharePointHelper(access_token)
            
            # Get configuration
            folder_url = self.sharepoint_config['folder_url']
            excel_file_name = self.sharepoint_config['excel_file_name']
            worksheet_name = self.sharepoint_config.get('worksheet_name', 'Sheet1')
            
            self.logger.info(f"Updating Excel attachment status for {request_id}")
            
            # Determine final attachment status
            attachment_status = 'Saved' if has_attachment else ('Failed' if attachment_error else 'None')
            has_attachment_text = 'Yes' if has_attachment else 'No'
            error_text = attachment_error if attachment_error else ''
            
            # Update Excel row
            sp.update_excel_row_by_id(
                folder_url=folder_url,
                excel_file_name=excel_file_name,
                worksheet_name=worksheet_name,
                id_column='A',
                id_value=request_id,
                updates={
                    'L': has_attachment_text,
                    'M': attachment_status,
                    'N': error_text
                }
            )
            
            self.logger.info(f"\033[92m✓ Excel attachment status updated for {request_id}\033[0m")
            return {'success': True}
            
        except Exception as e:
            self.logger.error(f"\033[91m✗ Failed to update Excel attachment status: {e}\033[0m")
            return {'success': False, 'error': str(e)}
