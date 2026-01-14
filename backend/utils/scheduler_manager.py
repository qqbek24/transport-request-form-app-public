"""
SchedulerManager - Orchestrates periodic background tasks
Encapsulates scheduler logic for JSON sync and attachment cleanup
"""
import json
import tempfile
import time
import logging
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

from sharepoint_helper import SharePointHelper

logger = logging.getLogger(__name__)


class SchedulerManager:
    """
    Manages scheduled background tasks
    Provides clean interface for JSON-to-SharePoint sync and attachment cleanup
    """
    
    def __init__(self, config: dict, transport_config: dict, 
                 get_access_token_func, app_logger_instance=None):
        """
        Initialize scheduler manager with dependencies
        
        Args:
            config: Full configuration dict
            transport_config: Transport section of config
            get_access_token_func: Function to get access token
            app_logger_instance: Structured app logger
        """
        self.config = config
        self.transport_config = transport_config
        self.sharepoint_config = transport_config.get('sharepoint', {})
        self.get_access_token = get_access_token_func
        self.app_logger = app_logger_instance
        self.logger = logger
    
    def sync_json_to_sharepoint(self):
        """
        Sync pending records from JSON backup to SharePoint Excel
        Runs periodically to retry failed SharePoint uploads
        """
        try:
            self.logger.info("\033[94mℹ Background sync: Starting JSON → SharePoint synchronization\033[0m")
            if self.app_logger:
                self.app_logger.log_info("Scheduled task: JSON to SharePoint sync started", {
                    'task': 'sync_json_to_sharepoint',
                    'trigger': 'scheduled'
                })
            
            # Check if SharePoint is enabled
            if not self.sharepoint_config.get('enabled', False):
                self.logger.info("\033[94mℹ Background sync: SharePoint integration disabled, skipping\033[0m")
                return
            
            # Get JSON path from config
            paths_config = self.transport_config.get('paths', {})
            json_path = Path(paths_config.get('json_backup_file', '/tmp/transport_requests.json'))
            
            if not json_path.exists():
                self.logger.info("\033[94mℹ Background sync: No JSON backup file found, nothing to sync\033[0m")
                return
            
            # Load JSON data
            with open(json_path, 'r', encoding='utf-8') as f:
                json_records = json.load(f)
            
            if not json_records:
                self.logger.info("\033[94mℹ Background sync: JSON backup file empty, nothing to sync\033[0m")
                return
            
            # Filter to only unsynced records
            unsynced_records = [r for r in json_records if not r.get('SharePoint_Synced', False)]
            
            self.logger.info(f"\033[94mℹ Background sync: Found {len(json_records)} total records, {len(unsynced_records)} unsynced\033[0m")
            
            if not unsynced_records:
                self.logger.info("\033[92m✓ Background sync: All records already synced to SharePoint\033[0m")
                return
            
            # Get access token and initialize SharePoint helper
            access_token = self.get_access_token()
            sp = SharePointHelper(access_token)
            
            folder_url = self.sharepoint_config['folder_url']
            excel_file_name = self.sharepoint_config['excel_file_name']
            worksheet_name = self.sharepoint_config.get('worksheet_name', 'Sheet1')
            
            # Download Excel to check which records already exist
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_path = Path(temp_dir) / excel_file_name
                
                self.logger.info(f"\033[94mℹ Background sync: Downloading {excel_file_name}\033[0m")
                folder = sp.get_folder(folder_url)
                sp.download_file_from_folder(
                    download_path=temp_dir,
                    folder=folder,
                    file_name=excel_file_name
                )
                
                wb = load_workbook(temp_path)
                ws = wb[worksheet_name] if worksheet_name in wb.sheetnames else wb.active
                
                # Get existing Request IDs from Excel (column A)
                existing_request_ids = set()
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                    cell_value = row[0].value
                    if cell_value and str(cell_value).strip():
                        existing_request_ids.add(str(cell_value))
                
                self.logger.info(f"\033[94mℹ Background sync: Found {len(existing_request_ids)} existing records in SharePoint Excel\033[0m")
                
                # Find unsynced records that are not in Excel
                records_to_sync = []
                for record in unsynced_records:
                    request_id = record.get('Request_ID')
                    if request_id and request_id not in existing_request_ids:
                        records_to_sync.append(record)
                
                if not records_to_sync:
                    self.logger.info("\033[92m✓ Background sync: All unsynced records already in SharePoint Excel\033[0m")
                    # Mark them as synced in JSON
                    self._mark_records_as_synced(json_path, json_records, unsynced_records, existing_request_ids)
                    return
                
                self.logger.info(f"\033[94mℹ Background sync: {len(records_to_sync)} records need to be synced\033[0m")
                
                # Sync missing records
                synced_count = 0
                failed_count = 0
                
                use_excel_api = self.sharepoint_config.get('use_excel_api', True)
                
                for record in records_to_sync:
                    try:
                        request_id = record.get('Request_ID')
                        
                        # Convert JSON record back to form data format
                        form_data = self._convert_record_to_form_data(record)
                        
                        has_attachment = record.get('Has_Attachment', 'No') == 'Yes'
                        attachment_error = record.get('Attachment_Error', '')
                        
                        # Save to SharePoint
                        if use_excel_api:
                            self._save_via_excel_api(request_id, form_data, has_attachment, attachment_error)
                        else:
                            self._save_via_traditional(request_id, form_data, has_attachment, attachment_error)
                        
                        synced_count += 1
                        self.logger.info(f"\033[92m✓ Background sync: Synced record {request_id} ({synced_count}/{len(records_to_sync)})\033[0m")
                        
                        # Mark as synced in JSON
                        record_json_index = json_records.index(record)
                        self._update_json_sync_status(json_path, json_records, record_json_index, True)
                        
                    except Exception as sync_error:
                        failed_count += 1
                        self.logger.error(f"\033[91m✗ Background sync: Failed to sync record {request_id}: {sync_error}\033[0m")
                
                self.logger.info(f"\033[94mℹ Background sync: Complete - {synced_count} synced, {failed_count} failed\033[0m")
                
        except Exception as e:
            self.logger.error(f"\033[91m✗ Background sync: Error during synchronization: {e}\033[0m", exc_info=True)
    
    def cleanup_old_attachments(self):
        """
        Delete attachments from SharePoint older than retention period
        Runs periodically to keep storage clean
        """
        try:
            self.logger.info("\033[94mℹ Attachment cleanup: Starting deletion of old files\033[0m")
            if self.app_logger:
                self.app_logger.log_info("Scheduled task: Attachment cleanup started", {
                    'task': 'cleanup_old_attachments',
                    'trigger': 'scheduled'
                })
            
            # Check if SharePoint is enabled
            if not self.sharepoint_config.get('enabled', False):
                self.logger.info("\033[94mℹ Attachment cleanup: SharePoint integration disabled, skipping\033[0m")
                return
            
            # Get retention period from config (default: 90 days = 3 months)
            retention_days = self.sharepoint_config.get('attachment_retention_days', 90)
            
            # Get access token
            access_token = self.get_access_token()
            sp = SharePointHelper(access_token)
            
            # Get attachments folder URL from config
            folder_url = self.sharepoint_config.get('folder_url')
            if not folder_url:
                self.logger.warning("\033[93m⚠ Attachment cleanup: No folder_url configured\033[0m")
                return
            
            # Assume attachments are in subfolder "Attachments"
            attachments_folder_url = f"{folder_url}/Attachments"
            
            try:
                attachments_folder = sp.get_folder(attachments_folder_url)
            except Exception as e:
                self.logger.warning(f"\033[93m⚠ Attachment cleanup: Attachments folder not found or error: {e}\033[0m")
                return
            
            # Get old files
            old_files = sp.get_files_older_than(attachments_folder, days=retention_days)
            
            if not old_files:
                self.logger.info(f"\033[94mℹ Attachment cleanup: No files older than {retention_days} days found\033[0m")
                return
            
            # Delete old files
            deleted_count = 0
            failed_count = 0
            
            for file_item in old_files:
                try:
                    if sp.delete_file(file_item):
                        deleted_count += 1
                        self.logger.info(f"\033[92m✓ Deleted: {file_item.get('name')}\033[0m")
                    else:
                        failed_count += 1
                except Exception as delete_error:
                    failed_count += 1
                    self.logger.error(f"\033[91m✗ Failed to delete {file_item.get('name')}: {delete_error}\033[0m")
            
            self.logger.info(f"\033[92m✓ Attachment cleanup: Complete - {deleted_count} deleted, {failed_count} failed\033[0m")
            
        except Exception as e:
            self.logger.error(f"\033[91m✗ Attachment cleanup: Error during cleanup: {e}\033[0m", exc_info=True)
    
    # ========== Private Helper Methods ==========
    
    def _convert_record_to_form_data(self, record: dict) -> dict:
        """Convert JSON record back to form data format"""
        return {
            'deliveryNoteNumber': record.get('Delivery_Note_Number', ''),
            'truckLicensePlates': record.get('Truck_License_Plates', ''),
            'trailerLicensePlates': record.get('Trailer_License_Plates', ''),
            'carrierCountry': record.get('Carrier_Country', ''),
            'carrierTaxCode': record.get('Carrier_Tax_Code', ''),
            'carrierFullName': record.get('Carrier_Full_Name', ''),
            'borderCrossing': record.get('Border_Crossing', ''),
            'borderCrossingDate': record.get('Border_Crossing_Date', ''),
            'email': record.get('Email', ''),
            'phoneNumber': record.get('Phone_Number', '')
        }
    
    def _save_via_excel_api(self, request_id: str, form_data: dict, 
                           has_attachment: bool, attachment_error: str):
        """
        Save record using Excel API (direct row insertion)
        Works even when file is open by other users
        """
        try:
            access_token = self.get_access_token()
            sp = SharePointHelper(access_token)
            
            folder_url = self.sharepoint_config['folder_url']
            excel_file_name = self.sharepoint_config['excel_file_name']
            worksheet_name = self.sharepoint_config.get('worksheet_name', 'Sheet1')
            
            self.logger.info(f"SharePoint Excel API: Adding row to {excel_file_name}")
            
            # Prepare row values in correct column order
            attachment_status = 'Saved' if has_attachment else ('Failed' if attachment_error else 'None')
            
            row_values = [
                request_id,
                form_data.get('deliveryNoteNumber', ''),
                form_data.get('truckLicensePlates', ''),
                form_data.get('trailerLicensePlates', ''),
                form_data.get('carrierCountry', ''),
                form_data.get('carrierTaxCode', ''),
                form_data.get('carrierFullName', ''),
                form_data.get('borderCrossing', ''),
                form_data.get('borderCrossingDate', ''),
                form_data.get('email', ''),
                form_data.get('phoneNumber', ''),
                'Yes' if has_attachment else 'No',
                attachment_status,
                attachment_error if attachment_error else ''
            ]
            
            # Add row using Excel API
            sp.add_excel_row(
                folder_url=folder_url,
                excel_file_name=excel_file_name,
                worksheet_name=worksheet_name,
                row_values=row_values
            )
            
            self.logger.info(f"\033[92m✓ SharePoint Excel API: Successfully added row for {request_id}\033[0m")
            
        except Exception as e:
            self.logger.error(f"\033[91m✗ SharePoint Excel API error: {e}\033[0m", exc_info=True)
            raise
    
    def _save_via_traditional(self, request_id: str, form_data: dict,
                             has_attachment: bool, attachment_error: str):
        """
        Save record using traditional download/upload method
        Uses retry logic for locked files
        """
        access_token = self.get_access_token()
        sp = SharePointHelper(access_token)
        
        folder_url = self.sharepoint_config['folder_url']
        excel_file_name = self.sharepoint_config['excel_file_name']
        worksheet_name = self.sharepoint_config.get('worksheet_name', 'Sheet1')
        max_retries = self.sharepoint_config.get('max_retries', 3)
        retry_wait_multiplier = self.sharepoint_config.get('retry_wait_multiplier', 2)
        
        self.logger.info(f"SharePoint: Getting folder from {folder_url}")
        folder = sp.get_folder(folder_url)
        
        # Retry loop for handling locked files
        last_error = None
        for attempt in range(max_retries):
            try:
                with tempfile.TemporaryDirectory() as temp_dir:
                    temp_path = Path(temp_dir) / excel_file_name
                    
                    self.logger.info(f"SharePoint: Downloading {excel_file_name} (attempt {attempt + 1}/{max_retries})")
                    sp.download_file_from_folder(
                        download_path=temp_dir,
                        folder=folder,
                        file_name=excel_file_name
                    )
                    
                    wb = load_workbook(temp_path)
                    ws = wb[worksheet_name] if worksheet_name in wb.sheetnames else wb.active
                    
                    # Prepare new row with data
                    new_row = [
                        request_id,
                        datetime.now().isoformat(),
                        form_data.get('deliveryNoteNumber', ''),
                        form_data.get('truckLicensePlates', ''),
                        form_data.get('trailerLicensePlates', ''),
                        form_data.get('carrierCountry', ''),
                        form_data.get('carrierTaxCode', ''),
                        form_data.get('carrierFullName', ''),
                        form_data.get('borderCrossing', ''),
                        form_data.get('borderCrossingDate', ''),
                        form_data.get('email', ''),
                        form_data.get('phoneNumber', ''),
                        'Yes' if has_attachment else 'No',
                        'Saved' if has_attachment else ('Failed' if attachment_error else 'None'),
                        attachment_error if attachment_error else ''
                    ]
                    
                    ws.append(new_row)
                    self.logger.info(f"SharePoint: Added new row {ws.max_row}")
                    
                    wb.save(temp_path)
                    wb.close()
                    
                    self.logger.info(f"SharePoint: Uploading {excel_file_name} (attempt {attempt + 1}/{max_retries})")
                    sp.upload_file(temp_path, folder)
                    
                    self.logger.info(f"\033[92m✓ SharePoint: Successfully saved to {excel_file_name}\033[0m")
                    return
                    
            except Exception as e:
                last_error = e
                error_str = str(e).lower()
                
                is_locked = any(indicator in error_str for indicator in [
                    'locked', 'in use', 'being used', 'cannot access', '423',
                    'cobaltlockviolation', 'resourcelocked', 'file is open'
                ])
                
                if is_locked and attempt < max_retries - 1:
                    wait_time = (attempt + 1) * retry_wait_multiplier
                    self.logger.warning(f"\033[93m⚠ SharePoint: File locked. Waiting {wait_time}s before retry...\033[0m")
                    time.sleep(wait_time)
                    continue
                else:
                    if is_locked:
                        self.logger.error(f"\033[91m✗ SharePoint: File still locked after {max_retries} attempts\033[0m")
                    raise
        
        self.logger.error(f"\033[91m✗ SharePoint: All {max_retries} attempts failed\033[0m")
        raise Exception(f"Failed to save to SharePoint after {max_retries} attempts. Last error: {last_error}")
    
    def _mark_records_as_synced(self, json_path: Path, json_records: list, 
                                unsynced_records: list, existing_request_ids: set):
        """Mark records as synced in JSON if they already exist in SharePoint"""
        for record in unsynced_records:
            if record.get('Request_ID') in existing_request_ids:
                record_json_index = json_records.index(record)
                self._update_json_sync_status(json_path, json_records, record_json_index, True)
    
    def _update_json_sync_status(self, json_path: Path, json_records: list, 
                                 record_index: int, synced: bool):
        """Update sync status for a specific record in JSON"""
        try:
            if 0 <= record_index < len(json_records):
                json_records[record_index]['SharePoint_Synced'] = synced
                
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(json_records, f, indent=2, ensure_ascii=False)
                    
        except Exception as e:
            self.logger.error(f"Failed to update JSON sync status: {e}")
