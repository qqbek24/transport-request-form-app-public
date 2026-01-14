"""
Test script dla SharePoint Helper
Sprawdza połączenie i podstawowe operacje
"""

import pytest
import os
import sys
from unittest.mock import Mock, patch

# Add backend to path
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'backend'))
from sharepoint_helper import SharePointHelper
from pathlib import Path
import os

# Wczytaj zmienne z pliku .env (jeśli istnieje)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv nie zainstalowane - użyj zmiennych systemowych

# KONFIGURACJA
# ------------
# Token z Azure - pobierany ze zmiennej środowiskowej
ACCESS_TOKEN = os.getenv('SHAREPOINT_ACCESS_TOKEN', 'WKLEJ_TUTAJ_TOKEN')

# URL folderu SharePoint (Twój link)
SHAREPOINT_FOLDER_URL = "https://yourcompany.sharepoint.com/sites/transport-test/Shared Documents/General/All Documents/transport-app"

# Nazwa pliku Excel
EXCEL_FILE_NAME = "transport_requests.xlsx"


def test_connection():
    """Test 1: Sprawdź połączenie z SharePoint"""
    print("\n" + "="*60)
    print("TEST 1: Połączenie z SharePoint")
    print("="*60)
    
    try:
        sp = SharePointHelper(ACCESS_TOKEN)
        folder = sp.get_folder(SHAREPOINT_FOLDER_URL)
        
        print(f"✅ Połączono z folderem: {folder['name']}")
        print(f"   Folder ID: {folder['id']}")
        print(f"   Web URL: {folder['webUrl']}")
        return sp, folder
        
    except Exception as e:
        print(f"❌ Błąd połączenia: {e}")
        return None, None


def test_list_files(sp, folder):
    """Test 2: Lista plików w folderze"""
    print("\n" + "="*60)
    print("TEST 2: Lista plików w folderze")
    print("="*60)
    
    try:
        children = sp.get_folder_childrens(folder)
        
        print(f"✅ Znaleziono {len(children)} elementów:")
        for child in children:
            item_type = "📁 Folder" if 'folder' in child else "📄 Plik"
            print(f"   {item_type}: {child['name']}")
        
        return children
        
    except Exception as e:
        print(f"❌ Błąd listowania: {e}")
        return []


def test_file_exists(sp, folder):
    """Test 3: Sprawdź czy plik Excel istnieje"""
    print("\n" + "="*60)
    print("TEST 3: Sprawdzanie istnienia pliku Excel")
    print("="*60)
    
    try:
        exists = sp.is_file_exists(folder, EXCEL_FILE_NAME)
        
        if exists:
            print(f"✅ Plik '{EXCEL_FILE_NAME}' istnieje w folderze")
        else:
            print(f"❌ Plik '{EXCEL_FILE_NAME}' NIE istnieje w folderze")
        
        return exists
        
    except Exception as e:
        print(f"❌ Błąd sprawdzania: {e}")
        return False


def test_download_file(sp, folder):
    """Test 4: Pobierz plik Excel"""
    print("\n" + "="*60)
    print("TEST 4: Pobieranie pliku Excel")
    print("="*60)
    
    try:
        # Pobierz do folderu temp
        temp_dir = Path(__file__).parent / "temp"
        temp_dir.mkdir(exist_ok=True)
        
        local_path = sp.download_file_from_folder(
            download_path=str(temp_dir),
            folder=folder,
            file_name=EXCEL_FILE_NAME
        )
        
        print(f"✅ Plik pobrany do: {local_path}")
        print(f"   Rozmiar: {local_path.stat().st_size} bajtów")
        
        return local_path
        
    except Exception as e:
        print(f"❌ Błąd pobierania: {e}")
        return None


def test_read_excel(file_path):
    """Test 5: Odczytaj plik Excel"""
    print("\n" + "="*60)
    print("TEST 5: Odczyt pliku Excel (openpyxl)")
    print("="*60)
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"✅ Plik Excel otwarty")
        print(f"   Nazwa arkusza: {ws.title}")
        print(f"   Wymiary: {ws.max_row} wierszy x {ws.max_column} kolumn")
        
        # Wyświetl pierwsze 3 wiersze
        print(f"\n   Pierwsze wiersze:")
        for row_idx in range(1, min(4, ws.max_row + 1)):
            row_data = [cell.value for cell in ws[row_idx]]
            print(f"   Row {row_idx}: {row_data[:5]}...")  # Pierwsze 5 kolumn
        
        wb.close()
        return True
        
    except ImportError:
        print("⚠️  Brak biblioteki openpyxl - zainstaluj: pip install openpyxl")
        return False
    except Exception as e:
        print(f"❌ Błąd odczytu: {e}")
        return False


def test_add_row_and_upload(sp, folder, file_path):
    """Test 6: Dodaj wiersz do Excel i upload"""
    print("\n" + "="*60)
    print("TEST 6: Dodaj wiersz i upload")
    print("="*60)
    
    try:
        from openpyxl import load_workbook
        from datetime import datetime
        
        # Otwórz plik
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Dodaj testowy wiersz
        test_row = [
            f"TEST-{datetime.now().strftime('%Y%m%d-%H%M%S')}",  # Request_ID
            datetime.now().isoformat(),  # Timestamp
            "TEST Delivery Note",
            "TEST-TRUCK-123",
            "TEST-TRAILER-456",
            "Poland",
            "TEST-TAX-789",
            "Test Carrier Name",
            "Giurgiu",
            "2025-11-14",
            "No"
        ]
        
        last_row = ws.max_row
        ws.append(test_row)
        
        print(f"✅ Dodano testowy wiersz {last_row + 1}")
        print(f"   Data: {test_row[:3]}")
        
        # Zapisz lokalnie
        wb.save(file_path)
        wb.close()
        print(f"✅ Zapisano lokalnie: {file_path}")
        
        # Upload do SharePoint
        result = sp.upload_file(file_path, folder)
        print(f"✅ Uploadowano do SharePoint")
        print(f"   ID: {result.get('id', 'N/A')}")
        
        return True
        
    except ImportError:
        print("⚠️  Brak biblioteki openpyxl - pomiń ten test")
        return False
    except Exception as e:
        print(f"❌ Błąd: {e}")
        return False


def main():
    """Uruchom wszystkie testy"""
    print("\n" + "="*60)
    print("TESTY SharePoint Helper")
    print("="*60)
    
    if ACCESS_TOKEN == 'WKLEJ_TUTAJ_TOKEN':
        print("\n❌ BŁĄD: Ustaw ACCESS_TOKEN w pliku test_sharepoint.py")
        print("   Lub ustaw zmienną środowiskową: SHAREPOINT_ACCESS_TOKEN")
        return
    
    # Test 1: Połączenie
    sp, folder = test_connection()
    if not sp or not folder:
        print("\n❌ Przerwano testy - brak połączenia")
        return
    
    # Test 2: Lista plików
    test_list_files(sp, folder)
    
    # Test 3: Sprawdź plik Excel
    file_exists = test_file_exists(sp, folder)
    if not file_exists:
        print("\n❌ Przerwano testy - brak pliku Excel")
        return
    
    # Test 4: Pobierz plik
    local_file = test_download_file(sp, folder)
    if not local_file:
        print("\n❌ Przerwano testy - nie udało się pobrać pliku")
        return
    
    # Test 5: Odczytaj Excel
    test_read_excel(local_file)
    
    # Test 6: Dodaj wiersz i upload (opcjonalny - ZAKOMENTUJ JEŚLI NIE CHCESZ MODYFIKOWAĆ PLIKU)
    print("\n⚠️  TEST 6 ZMODYFIKUJE PLIK W SHAREPOINT!")
    response = input("   Czy kontynuować? (tak/nie): ")
    if response.lower() in ['tak', 't', 'yes', 'y']:
        test_add_row_and_upload(sp, folder, local_file)
    else:
        print("   Pominięto test 6")
    
    print("\n" + "="*60)
    print("TESTY ZAKOŃCZONE")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()
